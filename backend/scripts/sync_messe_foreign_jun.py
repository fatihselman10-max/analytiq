"""
Messe foreign customer sync — Haziran 2026
Karar: YENİ müşterileri ekle (~41) + MEVCUTLARIN kategorisini (customer_type) sarı kolon F'den güncelle.
Mevcut müşterilerin diğer alanlarına DOKUNMAZ. Dry-run varsayılan; --apply ile yazar.

Usage:
    python3 sync_messe_foreign_jun.py "/path/foreign customer1 (1) (1).xlsx"          # dry-run
    python3 sync_messe_foreign_jun.py "/path/foreign customer1 (1) (1).xlsx" --apply   # uygula
"""
import sys, re
import openpyxl
import psycopg2
from collections import Counter, defaultdict

DB = dict(host="junction.proxy.rlwy.net", port=51834, user="postgres",
          password="xswhBJfiUcWyTQdDdMvBGxXsMOkwJObO", dbname="railway")
ORG_ID = 1
AGENT_MAP = {"Meryem": 3, "Samet": 2}

COUNTRY_NORM = {
    "RUSYA":"Rusya","TUNUS":"Tunus","BELARUS":"Belarus","BELARUS ":"Belarus",
    "UKRAYNA":"Ukrayna","SIRBISTAN":"Sırbistan","MOLDOVA":"Moldova","İTALYA":"İtalya",
    "BULGARİSTAN":"Bulgaristan","İRAN":"İran","GÜRCİSTAN":"Gürcistan","MISIR":"Mısır",
    "ÇEKYA":"Çekya","LÜBNAN":"Lübnan","LİTVANYA":"Litvanya","POLONYA":"Polonya",
    "PORTEKİZ":"Portekiz","FAS":"Fas","ESTONYA":"Estonya","ALMANYA":"Almanya",
    "İngiltere":"İngiltere","ÖZBEKİSTAN":"Özbekistan",
}
SEG_TO_PIPELINE = {1:"order_received",2:"kartela_sent",3:"new_contact",4:"new_contact"}

def classify_activity(text):
    if not text: return None,None,None
    t=text.strip().lower()
    if "rusca tanıtım" in t or "rusça tanıtım" in t or "tanıtım film" in t or "tanitim film" in t:
        return "intro_video_sent","Rusça tanıtım filmi gönderildi",text
    if "depo video" in t: return "warehouse_video_sent","Depo videosu gönderildi",text
    if "moskova" in t and "minsk" in t: return "fair_invitation","Moskova-Minsk fuarına davet edildi",text
    if "fuar" in t and ("davet" in t or "gönder" in t): return "fair_invitation","Fuara davet edildi",text
    if "telegram" in t and "tanıt" in t: return "initial_contact","Telegram'dan firma tanıtıldı",text
    if "whatsapp" in t and "tanıt" in t: return "initial_contact","WhatsApp'tan firma tanıtıldı",text
    if "toplu" in t and ("mesaj" in t or "görsel" in t or "gorsel" in t): return "bulk_message","Toplu bilgilendirme / tanıtım gönderildi",text
    if "katalog" in t: return "catalog_request",text[:80],text
    if "kartela" in t: return "kartela_request",text[:80],text
    if "numune" in t: return "sample_request",text[:80],text
    if "sezon" in t: return "price_inquiry",text[:80],text
    return "note",text[:80],text

def norm_channel(s):
    if not s: return ""
    s=str(s).strip().upper()
    if "WHATSAPP" in s or "WHATS" in s: return "whatsapp"
    if "TELEGRAM" in s: return "telegram"
    if "INSTAGRAM" in s: return "instagram"
    if "MAIL" in s or "EMAIL" in s: return "email"
    if "VK" in s: return "vk"
    if "TELEFON" in s or "PHONE" in s or "ÇAĞRI" in s or "CAGRI" in s: return "phone"
    return s.lower()

def clean_phone(p):
    if not p: return ""
    digits=re.sub(r"[^\d+]","",str(p).strip())
    if digits.startswith("00"): digits="+"+digits[2:]
    elif not digits.startswith("+") and digits.startswith("7") and len(digits)==11: digits="+"+digits
    return digits[:50]

def cs(s,maxlen=None):
    if s is None: return ""
    s=str(s).strip()
    return s[:maxlen] if maxlen else s

def main(xlsx_path, apply=False):
    wb=openpyxl.load_workbook(xlsx_path, data_only=True)
    ws=wb['foreign customer']
    rows=[]
    for r in range(2, ws.max_row+1):
        row=[ws.cell(r,c).value for c in range(1,29)]
        if any(v for v in row[:10]): rows.append(row)
    print(f"Excel veri satırı: {len(rows)}")

    conn=psycopg2.connect(**DB); conn.autocommit=False; c=conn.cursor()
    # Mevcut müşteriler: lowercased name/company -> [(id, customer_type)]
    c.execute("SELECT id, lower(name), lower(COALESCE(company,'')), COALESCE(customer_type,'') FROM customers WHERE org_id=%s",(ORG_ID,))
    by_key=defaultdict(list)
    for cid,n,co,ct in c.fetchall():
        if n: by_key[n].append((cid,ct))
        if co and co!=n: by_key[co].append((cid,ct))

    new_count=cat_updated=cat_same=cat_empty=ambiguous=act_count=0
    new_preview=[]; upd_preview=[]

    for row in rows:
        country=COUNTRY_NORM.get(cs(row[0]),cs(row[0]))
        try:
            segment=int(row[1]);  segment = segment if segment in (1,2,3,4) else 4
        except (TypeError,ValueError): segment=4
        way=cs(row[2],255); responsible=cs(row[3]); brand=cs(row[4],255)
        category=cs(row[5],100); contact=cs(row[6],255)
        instagram=cs(row[7],255); website=cs(row[8],255); vk=cs(row[9],255)
        telegram=cs(row[10],255); phone=clean_phone(row[11]); email=cs(row[12],255)
        fabric=cs(row[13]); preferred=norm_channel(row[15])
        act1,act2,act3=row[16],row[17],row[18]; note=cs(row[19])
        name=contact or brand or f"{country} Müşteri"
        if not name: continue

        # Eşleşme ara
        ids=set()
        for k in {contact.lower(), brand.lower()}:
            if k:
                for cid,ct in by_key.get(k,[]): ids.add((cid,ct))
        ids=list(ids)

        if not ids:
            # YENİ müşteri
            new_count+=1
            assigned=AGENT_MAP.get(responsible)
            pipeline=SEG_TO_PIPELINE.get(segment,"new_contact")
            source="fair" if any(x in way.upper() for x in ["CPM","TS ","INTERTEX","B2B","VIPTEX","VIP TEX"]) else ("other" if way else "")
            new_preview.append(f"{name} | {brand} | {country} | S{segment} | {category}")
            if apply:
                c.execute("""INSERT INTO customers
                  (org_id,name,company,country,segment,customer_type,source,source_detail,assigned_to,
                   phone,email,instagram,website,vk,telegram,preferred_channel,notes,interested_products,
                   pipeline_stage,pipeline_updated_at,last_contact_at,created_at,updated_at)
                  VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),NOW(),NOW(),NOW())
                  RETURNING id""",
                  (ORG_ID,name,brand,country,segment,category,source,way,assigned,
                   phone,email,instagram,website,vk,telegram,preferred,note,fabric,pipeline))
                cid=c.fetchone()[0]
                for raw in (act1,act2,act3):
                    if not raw: continue
                    at,ati,ad=classify_activity(str(raw))
                    if not at: continue
                    c.execute("""INSERT INTO customer_activities
                      (org_id,customer_id,activity_type,title,description,status,detected_by,created_at)
                      VALUES (%s,%s,%s,%s,%s,'approved','manual',NOW())""",(ORG_ID,cid,at,ati,ad))
                    act_count+=1
        else:
            # MEVCUT — sadece kategori güncelle
            if not category:
                cat_empty+=1; continue
            if len(ids)>1:
                ambiguous+=1; continue
            cid,cur=ids[0]
            if cur.strip().lower()==category.strip().lower():
                cat_same+=1
            else:
                cat_updated+=1
                upd_preview.append(f"#{cid} {name}: '{cur}' -> '{category}'")
                if apply:
                    c.execute("UPDATE customers SET customer_type=%s, updated_at=NOW() WHERE id=%s AND org_id=%s",
                              (category,cid,ORG_ID))

    if apply: conn.commit()
    else: conn.rollback()
    conn.close()

    print(f"\n{'UYGULANDI' if apply else 'DRY-RUN'}")
    print(f"  Yeni müşteri eklendi : {new_count}  (+{act_count} aktivite)")
    print(f"  Kategori güncellendi : {cat_updated}")
    print(f"  Kategori zaten aynı  : {cat_same}")
    print(f"  Dosyada kategori boş : {cat_empty} (atlandı)")
    print(f"  Belirsiz (çok eşleşme): {ambiguous} (atlandı)")
    print("\n--- YENİ (ilk 50) ---")
    for x in new_preview[:50]: print("  +",x)
    print("\n--- KATEGORİ DEĞİŞİMİ ---")
    for x in upd_preview: print("  ~",x)

if __name__=="__main__":
    path=sys.argv[1]
    apply="--apply" in sys.argv
    main(path, apply)
