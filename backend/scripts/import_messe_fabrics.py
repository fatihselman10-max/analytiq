"""
Messe Tekstil Fabric Catalog Import
- Reads FABRIC LIST.xlsx (FW26 + SS26 sheets)
- Maps to folders under /tmp/fw26_fabrics and /tmp/ss26_fabrics
- Compresses PNG → JPG 85%
- Upserts fabrics + fabric_images rows
"""
import os, sys, io, re
import openpyxl
import psycopg2
from PIL import Image

DB = dict(host="junction.proxy.rlwy.net", port=51834, user="postgres",
          password="xswhBJfiUcWyTQdDdMvBGxXsMOkwJObO", dbname="railway")
ORG_ID = 1

FW26_DIR = "/tmp/fw26_fabrics"
SS26_DIR = "/tmp/ss26_fabrics"
XLSX = "/tmp/fabric_list.xlsx"

def compress_jpg(png_path, quality=85, max_w=1600):
    img = Image.open(png_path)
    if img.mode != "RGB":
        img = img.convert("RGB")
    if img.width > max_w:
        ratio = max_w / img.width
        img = img.resize((max_w, int(img.height * ratio)), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=quality, optimize=True)
    return buf.getvalue()

def find_folder(code, season_dir):
    # Try exact match, then prefix matches (handles 6775-3 → 6775)
    if os.path.isdir(os.path.join(season_dir, code)):
        return os.path.join(season_dir, code)
    base = code.split("-")[0]
    if os.path.isdir(os.path.join(season_dir, base)):
        return os.path.join(season_dir, base)
    # Try code with -1 / -2 suffix
    for entry in os.listdir(season_dir):
        if entry == code or entry.startswith(code + "-") or entry.startswith(base):
            if os.path.isdir(os.path.join(season_dir, entry)):
                return os.path.join(season_dir, entry)
    return None

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    conn = psycopg2.connect(**DB)
    conn.autocommit = True
    c = conn.cursor()

    total_fabrics = 0
    total_images = 0
    no_images = []
    missing_folder = []

    for sheet_name, season_dir in [("FW26", FW26_DIR), ("SS26", SS26_DIR)]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        print(f"\n=== {sheet_name} ===", flush=True)

        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            code = ws.cell(r, 2).value
            width = ws.cell(r, 3).value
            comp = ws.cell(r, 4).value
            gauge = ws.cell(r, 5).value
            if not code:
                continue
            code = str(code).strip()
            name = str(name or "").strip()

            # Upsert fabric
            c.execute("""
                INSERT INTO fabrics (org_id, code, name, season, width, composition, gauge)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (org_id, code) DO UPDATE SET
                  name=EXCLUDED.name, season=EXCLUDED.season,
                  width=EXCLUDED.width, composition=EXCLUDED.composition,
                  gauge=EXCLUDED.gauge, updated_at=NOW()
                RETURNING id
            """, (ORG_ID, code, name, sheet_name, str(width or ""), str(comp or ""), str(gauge or "")))
            fab_id = c.fetchone()[0]
            total_fabrics += 1

            # Find images folder
            folder = find_folder(code, season_dir)
            if not folder:
                missing_folder.append(f"{sheet_name}/{code}")
                print(f"  [no-folder] {code} {name}", flush=True)
                continue

            imgs = sorted([f for f in os.listdir(folder)
                           if f.lower().endswith((".png", ".jpg", ".jpeg", ".webp"))])
            if not imgs:
                no_images.append(f"{sheet_name}/{code}")
                continue

            # Clear existing to allow re-import
            c.execute("DELETE FROM fabric_images WHERE fabric_id=%s", (fab_id,))

            for i, fname in enumerate(imgs):
                fpath = os.path.join(folder, fname)
                try:
                    jpg_data = compress_jpg(fpath)
                except Exception as e:
                    print(f"  ! compress error {fname}: {e}", flush=True)
                    continue
                c.execute("""
                    INSERT INTO fabric_images (fabric_id, file_name, file_type, file_size, file_data, sort_order)
                    VALUES (%s,%s,'image/jpeg',%s,%s,%s)
                """, (fab_id, fname.replace(".png",".jpg").replace(".PNG",".jpg"),
                      len(jpg_data), psycopg2.Binary(jpg_data), i))
                total_images += 1
            print(f"  {code:8s} {name:20s} → {len(imgs)} görsel", flush=True)

    conn.close()
    print(f"\n✅ {total_fabrics} kumaş, {total_images} görsel yüklendi")
    if missing_folder:
        print(f"\n⚠ {len(missing_folder)} kumaş görseli YOK:")
        for m in missing_folder:
            print(f"  - {m}")

if __name__ == "__main__":
    main()
