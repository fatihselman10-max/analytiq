"""
Messe Tekstil Excel Import — Foreign Customers
Reads 'foreign customer1.xlsx' and inserts customers + activities into messe-backend DB.

Usage:
    python3 import_messe_foreign_customers.py /path/to/foreign\ customer1.xlsx
"""
import sys, re, json
import openpyxl
import psycopg2
from collections import Counter

DB = dict(host="junction.proxy.rlwy.net", port=51834, user="postgres",
          password="xswhBJfiUcWyTQdDdMvBGxXsMOkwJObO", dbname="railway")
ORG_ID = 1

# User mapping: Excel responsible → users.id
AGENT_MAP = {"Meryem": 3, "Samet": 2}  # Fatih=1, Samet=2, Meryem=3

# Country Turkish-normalize
COUNTRY_NORM = {
    "RUSYA":"Rusya", "TUNUS":"Tunus", "BELARUS":"Belarus", "BELARUS ":"Belarus",
    "UKRAYNA":"Ukrayna", "SIRBISTAN":"Sırbistan", "MOLDOVA":"Moldova",
    "İTALYA":"İtalya", "BULGARİSTAN":"Bulgaristan", "İRAN":"İran",
    "GÜRCİSTAN":"Gürcistan", "MISIR":"Mısır", "ÇEKYA":"Çekya",
    "LÜBNAN":"Lübnan", "LİTVANYA":"Litvanya", "POLONYA":"Polonya",
    "PORTEKİZ":"Portekiz", "FAS":"Fas", "ESTONYA":"Estonya",
    "ALMANYA":"Almanya", "İngiltere":"İngiltere", "ÖZBEKİSTAN":"Özbekistan",
}

# Segment → pipeline_stage
SEG_TO_PIPELINE = {
    1: "order_received",  # Sales + communication exist
    2: "kartela_sent",    # Communication but no sale
    3: "new_contact",     # Met at fair, big company, silent
    4: "new_contact",     # Met at fair, small, silent
}

# Normalized activity types — decide by keyword
def classify_activity(text):
    if not text: return None, None, None
    t = text.strip().lower()
    if "rusca tanıtım" in t or "rusça tanıtım" in t or "tanıtım film" in t or "tanitim film" in t:
        return "intro_video_sent", "Rusça tanıtım filmi gönderildi", text
    if "depo video" in t:
        return "warehouse_video_sent", "Depo videosu gönderildi", text
    if "moskova" in t and "minsk" in t:
        return "fair_invitation", "Moskova-Minsk fuarına davet edildi", text
    if "fuar" in t and ("davet" in t or "gönder" in t):
        return "fair_invitation", "Fuara davet edildi", text
    if "telegram" in t and "tanıt" in t:
        return "initial_contact", "Telegram'dan firma tanıtıldı", text
    if "whatsapp" in t and "tanıt" in t:
        return "initial_contact", "WhatsApp'tan firma tanıtıldı", text
    if "toplu" in t and ("mesaj" in t or "görsel" in t or "gorsel" in t):
        return "bulk_message", "Toplu bilgilendirme / tanıtım gönderildi", text
    if "katalog" in t:
        return "catalog_request", text[:80], text
    if "kartela" in t:
        return "kartela_request", text[:80], text
    if "numune" in t:
        return "sample_request", text[:80], text
    if "sezon" in t:
        return "price_inquiry", text[:80], text
    # Fallback - unknown
    return "note", text[:80], text

# Preferred channel normalize
def norm_channel(s):
    if not s: return ""
    s = str(s).strip().upper()
    if "WHATSAPP" in s or "WHATS" in s: return "whatsapp"
    if "TELEGRAM" in s: return "telegram"
    if "INSTAGRAM" in s: return "instagram"
    if "MAIL" in s or "EMAIL" in s: return "email"
    if "VK" in s: return "vk"
    if "TELEFON" in s or "PHONE" in s or "ÇAĞRI" in s or "CAGRI" in s: return "phone"
    return s.lower()

def clean_phone(p):
    if not p: return ""
    p = str(p).strip()
    # Keep digits and + only; normalize leading 7 (Russia) to +7
    digits = re.sub(r"[^\d+]", "", p)
    if digits.startswith("00"): digits = "+" + digits[2:]
    elif not digits.startswith("+") and digits.startswith("7") and len(digits) == 11:
        digits = "+" + digits
    elif not digits.startswith("+"):
        digits = digits
    return digits[:50]

def clean_str(s, maxlen=None):
    if s is None: return ""
    s = str(s).strip()
    if maxlen: s = s[:maxlen]
    return s

def main(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['foreign customer']

    rows = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 29)]
        if any(v for v in row[:10]):
            rows.append(row)

    print(f"Reading {len(rows)} rows...")

    conn = psycopg2.connect(**DB)
    conn.autocommit = True
    c = conn.cursor()

    imported = 0
    act_count = 0
    skipped = 0
    seg_dist = Counter()
    act_type_dist = Counter()

    for row in rows:
        country = COUNTRY_NORM.get(clean_str(row[0]), clean_str(row[0]))
        try:
            segment = int(row[1])
            if segment not in (1,2,3,4): segment = 4
        except (TypeError, ValueError):
            segment = 4
        way_to_reach = clean_str(row[2], 255)   # source_detail (CPM, TS 2026, VIPTEX…)
        responsible = clean_str(row[3])          # Meryem/Samet
        brand = clean_str(row[4], 255)           # company
        category = clean_str(row[5], 100)        # customer_type
        contact_person = clean_str(row[6], 255)  # name
        instagram = clean_str(row[7], 255)
        website = clean_str(row[8], 255)
        vk = clean_str(row[9], 255)
        telegram = clean_str(row[10], 255)
        phone = clean_phone(row[11])
        email = clean_str(row[12], 255)
        fabric = clean_str(row[13])              # interested_products
        # row[14] fabricsell (mostly empty)
        preferred = norm_channel(row[15])
        act1 = row[16]
        act2 = row[17]
        act3 = row[18]
        note = clean_str(row[19])

        # Name fallback: if contact_person empty use brand
        name = contact_person or brand or f"{country} Müşteri"
        if not name:
            skipped += 1
            continue

        assigned_to = AGENT_MAP.get(responsible)
        pipeline = SEG_TO_PIPELINE.get(segment, "new_contact")
        source_channel = ""  # Excel column (VIPTEX/CPM/TS…) is detail, not channel; source=fair
        source = "fair" if any(x in way_to_reach.upper() for x in ["CPM","TS ","INTERTEX","B2B","VIPTEX","VIP TEX"]) else ""
        if not source and way_to_reach:
            source = "other"

        # Insert customer
        c.execute("""
            INSERT INTO customers
              (org_id, name, company, country, segment, customer_type,
               source, source_detail, assigned_to, phone, email, instagram,
               website, vk, telegram, preferred_channel,
               notes, interested_products, pipeline_stage, pipeline_updated_at,
               last_contact_at, created_at, updated_at)
            VALUES (%s,%s,%s,%s,%s,%s, %s,%s,%s,%s,%s,%s, %s,%s,%s,%s,
                    %s,%s,%s,NOW(), NOW(),NOW(),NOW())
            RETURNING id
        """, (ORG_ID, name, brand, country, segment, category,
              source, way_to_reach, assigned_to, phone, email, instagram,
              website, vk, telegram, preferred,
              note, fabric, pipeline))
        cust_id = c.fetchone()[0]
        imported += 1
        seg_dist[segment] += 1

        # Activities — approved (already manually done by the team)
        for raw in (act1, act2, act3):
            if not raw: continue
            atype, atitle, adesc = classify_activity(str(raw))
            if not atype: continue
            c.execute("""
                INSERT INTO customer_activities
                  (org_id, customer_id, activity_type, title, description,
                   status, detected_by, created_at)
                VALUES (%s,%s,%s,%s,%s,'approved','manual',NOW())
            """, (ORG_ID, cust_id, atype, atitle, adesc))
            act_count += 1
            act_type_dist[atype] += 1

        if imported % 20 == 0:
            print(f"  [{imported}/{len(rows)}] kayıt edildi...", flush=True)

    conn.close()

    print(f"\n✅ Imported {imported} customers ({skipped} skipped)")
    print(f"✅ Imported {act_count} activities")
    print("\nSegment dağılımı:")
    for s, n in sorted(seg_dist.items()):
        print(f"  {s}: {n}")
    print("\nAktivite tip dağılımı:")
    for t, n in act_type_dist.most_common():
        print(f"  {t:25s} {n}")

if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "/Users/fatih/Downloads/foreign customer1.xlsx"
    main(path)
